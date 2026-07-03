"""秘密度ラベル (MSIP) と暗号化/IRM 保護の検知・伝播を検証する。

- ``detect_protection``: 暗号化/IRM を検知し、通常ファイルは誤検知しない
- ``read_label``: MSIP ラベルを custom.xml から読む
- ``extract()``: 保護文書は ``ProtectedDocumentError`` で fail-closed、
  ラベルは result.json の ``metadata.sensitivity`` と index.json へ伝播する
"""

from __future__ import annotations

import json
import sys
import zipfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import pytest

from docextract import extract, sensitivity
from docextract.sensitivity import ProtectedDocumentError

_OLE = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
_GUID = "2096f6a2-d2f7-48be-b329-b73aaa526e5d"


def _u16(s: str) -> bytes:
    return s.encode("utf-16-le")


def _custom_xml(name: str, enabled: str = "True") -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/custom-properties"'
        ' xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">'
        f'<property fmtid="{{X}}" pid="2" name="MSIP_Label_{_GUID}_Enabled"><vt:lpwstr>{enabled}</vt:lpwstr></property>'
        f'<property fmtid="{{X}}" pid="3" name="MSIP_Label_{_GUID}_Name"><vt:lpwstr>{name}</vt:lpwstr></property>'
        f'<property fmtid="{{X}}" pid="4" name="MSIP_Label_{_GUID}_SetDate"><vt:lpwstr>2026-07-01T00:00:00Z</vt:lpwstr></property>'
        f'<property fmtid="{{X}}" pid="5" name="MSIP_Label_{_GUID}_Method"><vt:lpwstr>Standard</vt:lpwstr></property>'
        "</Properties>"
    )


def _labeled_xlsx(path: Path, name: str = "社外秘") -> Path:
    """openpyxl で作った実 xlsx に MSIP ラベルの custom.xml を注入する。"""
    from openpyxl import Workbook

    wb = Workbook()
    wb.active["A1"] = "本文セル"
    wb.save(path)
    # zip に docProps/custom.xml を追記 (read_label はこのパートを読む)
    with zipfile.ZipFile(path, "a") as z:
        z.writestr("docProps/custom.xml", _custom_xml(name))
    return path


# --- detect_protection ------------------------------------------------------


def test_normal_ooxml_is_not_flagged(tmp_path):
    p = tmp_path / "ok.docx"
    with zipfile.ZipFile(p, "w") as z:
        z.writestr("word/document.xml", "<x/>")
    assert sensitivity.detect_protection(p) is None


def test_normal_legacy_ole_is_not_flagged(tmp_path):
    """通常の (無保護の) 旧 Office 文書は OLE2 だが保護構造を持たない。"""
    p = tmp_path / "plain.doc"
    p.write_bytes(_OLE + b"WordDocument" * 20)
    assert sensitivity.detect_protection(p) is None


def test_irm_protected_is_detected(tmp_path):
    p = tmp_path / "secret.doc"
    p.write_bytes(_OLE + b"..." + _u16("DRMEncryptedTransform") + b"...")
    info = sensitivity.detect_protection(p)
    assert info is not None and info["kind"] == "irm"


def test_password_encrypted_is_detected(tmp_path):
    p = tmp_path / "enc.xlsx"
    p.write_bytes(_OLE + _u16("EncryptedPackage"))
    info = sensitivity.detect_protection(p)
    assert info is not None and info["kind"] == "encrypted"


def test_marker_across_chunk_boundary(tmp_path):
    """マーカーがチャンク境界をまたいでも検知できる (取りこぼさない)。"""
    marker = _u16("EncryptedPackage")
    chunk = 1 << 20
    p = tmp_path / "big.xls"
    # OLE ヘッダ直後～境界直前まで詰め、境界にマーカーがまたがるよう配置
    pad = b"\x00" * (chunk - len(_OLE) - len(marker) // 2)
    p.write_bytes(_OLE + pad + marker + b"\x00" * 10)
    assert sensitivity.detect_protection(p)["kind"] == "encrypted"


# --- read_label -------------------------------------------------------------


def test_read_label_parses_msip(tmp_path):
    p = tmp_path / "labeled.xlsx"
    with zipfile.ZipFile(p, "w") as z:
        z.writestr("docProps/custom.xml", _custom_xml("社外秘"))
    label = sensitivity.read_label(p)
    assert label["name"] == "社外秘"
    assert label["id"] == _GUID
    assert label["enabled"] is True
    assert label["method"] == "Standard"


def test_read_label_absent_returns_none(tmp_path):
    p = tmp_path / "nolabel.xlsx"
    with zipfile.ZipFile(p, "w") as z:
        z.writestr("docProps/core.xml", "<x/>")
    assert sensitivity.read_label(p) is None


def test_read_label_ignores_non_zip(tmp_path):
    p = tmp_path / "legacy.doc"
    p.write_bytes(_OLE + b"whatever")
    assert sensitivity.read_label(p) is None


# --- extract() 統合 ---------------------------------------------------------


def test_extract_rejects_password_encrypted(tmp_path):
    """パスワード暗号化はアクセス権とは別に鍵が要るため fail-closed する。"""
    p = tmp_path / "secret.xlsx"
    p.write_bytes(_OLE + _u16("EncryptedPackage"))  # DRM マーカー無し = パスワード暗号化
    with pytest.raises(ProtectedDocumentError) as ei:
        extract(p, output_dir=tmp_path / "out", ocr=False, image_tables=False)
    assert "暗号化" in str(ei.value)
    # 復号平文の成果物を書き出していない
    assert not (tmp_path / "out").exists() or not any(
        (tmp_path / "out").glob("*/result.json")
    )


def test_extract_irm_routes_to_office_decrypt(tmp_path):
    """IRM/RMS はブロックせず Office 復号経路へ。Office 不在では Office 必須で停止。"""
    from docextract.extractors import OfficeUnavailableError

    p = tmp_path / "secret.docx"
    p.write_bytes(_OLE + _u16("DRMEncryptedTransform"))
    # CI には Office/pywin32 が無いので、復号経路に入り OfficeUnavailableError になる
    # (= ProtectedDocumentError での門前払いにはならない = ブロックしていない)
    with pytest.raises(OfficeUnavailableError) as ei:
        extract(p, output_dir=tmp_path / "out", ocr=False, image_tables=False)
    msg = str(ei.value)
    assert "Office" in msg and "復号" in msg
    assert not any((tmp_path / "out").glob("*/result.json"))


def test_extract_irm_decrypts_and_extracts_with_office(tmp_path, monkeypatch):
    """Office が使える場合、IRM 文書は復号され通常どおり抽出される（モック COM）。"""
    from docextract.extractors import legacy_com

    monkeypatch.setattr(legacy_com, "_require_win32com", lambda ext, app, action: None)

    # 復号後の平文 OOXML を模す実 xlsx を用意し、変換ダミーがそれをコピーする
    from openpyxl import Workbook

    seed = tmp_path / "decrypted_seed.xlsx"
    wb = Workbook()
    wb.active["A1"] = "復号後の本文"
    wb.save(seed)

    def _fake_convert(src, dst):
        import shutil

        shutil.copyfile(seed, dst)

    monkeypatch.setitem(legacy_com._APP_SPEC["Excel"], "convert", _fake_convert)

    p = tmp_path / "secret.xlsx"
    p.write_bytes(_OLE + _u16("DRMEncryptedTransform"))  # irm 判定させる
    data = extract(p, output_dir=tmp_path / "out", ocr=False, image_tables=False)

    assert data["file_type"] == "xlsx"
    assert "COM decrypt" in data["metadata"]["converted_via"]
    # 孤立した単一セルはテキスト要素として出る
    cells = [
        el.get("content") for el in data["elements"] if el["type"] == "text"
    ] + [
        c
        for el in data["elements"]
        if el["type"] == "table"
        for row in el["rows"]
        for c in row
    ]
    assert "復号後の本文" in cells


def test_extract_propagates_label_to_result_and_manifest(tmp_path):
    src = _labeled_xlsx(tmp_path / "report.xlsx", name="社外秘")
    out = tmp_path / "out"
    data = extract(src, output_dir=out, ocr=False, image_tables=False)

    # result.json (返り値) にラベルが載る
    sens = data["metadata"]["sensitivity"]
    assert sens["name"] == "社外秘"
    assert sens["id"] == _GUID

    # index.json (マニフェスト) にもラベルが載る
    index = json.loads((out / "index.json").read_text(encoding="utf-8"))
    entries = index["documents"] if isinstance(index, dict) else index
    rec = next(iter(entries.values())) if isinstance(entries, dict) else entries[0]
    assert rec["sensitivity"]["name"] == "社外秘"


def test_extract_without_label_has_no_sensitivity(tmp_path):
    from openpyxl import Workbook

    p = tmp_path / "plain.xlsx"
    wb = Workbook()
    wb.active["A1"] = "x"
    wb.save(p)
    data = extract(p, output_dir=tmp_path / "out", ocr=False, image_tables=False)
    assert "sensitivity" not in data["metadata"]
